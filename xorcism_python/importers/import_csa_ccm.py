"""import_csa_ccm.py — import the CSA Cloud Controls Matrix (CCM) into XORCISM.CONTROL.

The Cloud Security Alliance CCM v4 is the de-facto cloud-security control framework: 197
controls across 17 domains, each with a CAIQ assessment question and rich mappings to ISO
27001/27017/27018, NIST 800-53, PCI DSS, SOC 2, etc.

Two modes:
  * --file <CCM.xlsx>  : parse the official CSA spreadsheet (openpyxl) — all controls + CAIQ +
                         the Control ID/Title/Specification columns. (Download from cloudsecurityalliance.org.)
  * (no file)          : seed the 17 domains + a representative set of key controls (embedded),
                         so the framework is present immediately; run --file later for the full 197.

Writes CONTROL rows (VocabularyID = the CCM vocab; ControlName="CCM-ID Title", CIS-style domain in
ControlDescription, Statement=specification) keyed idempotently by VocabularyID. Registers the
VOCABULARY entry "CSA CCM". Raw SQL; DB = XORCISM_DB_DIR env or default.

    python xorcism_python/importers/import_csa_ccm.py [--file CCM_v4.xlsx]
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import uuid
from datetime import datetime, timezone

VOCAB = "CSA CCM"

# 17 CCM v4 domains (id -> name).
DOMAINS = {
    "A&A": "Audit & Assurance", "AIS": "Application & Interface Security",
    "BCR": "Business Continuity Mgmt & Operational Resilience", "CCC": "Change Control & Configuration Management",
    "CEK": "Cryptography, Encryption & Key Management", "DCS": "Datacenter Security",
    "DSP": "Data Security & Privacy Lifecycle Management", "GRC": "Governance, Risk & Compliance",
    "HRS": "Human Resources Security", "IAM": "Identity & Access Management",
    "IPY": "Interoperability & Portability", "IVS": "Infrastructure & Virtualization Security",
    "LOG": "Logging & Monitoring", "SEF": "Security Incident Mgmt, E-Discovery & Cloud Forensics",
    "STA": "Supply Chain Mgmt, Transparency & Accountability", "TVM": "Threat & Vulnerability Management",
    "UEM": "Universal Endpoint Management",
}

# Representative key controls (id, title). The full 197 come from --file.
SEED = [
    ("A&A-01", "Audit and Assurance Policy and Procedures"),
    ("A&A-02", "Independent Assessments"),
    ("AIS-01", "Application and Interface Security Policy and Procedures"),
    ("AIS-04", "Secure Application Design and Development"),
    ("AIS-07", "Application Vulnerability Remediation"),
    ("BCR-01", "Business Continuity Management Policy and Procedures"),
    ("BCR-08", "Backup"),
    ("CCC-01", "Change Management Policy and Procedures"),
    ("CCC-06", "Change Management Baseline"),
    ("CEK-01", "Encryption and Key Management Policy and Procedures"),
    ("CEK-03", "Data Encryption"),
    ("CEK-09", "Encryption Key Rotation"),
    ("DCS-01", "Off-Site Equipment Disposal Policy and Procedures"),
    ("DCS-06", "Assets Classification"),
    ("DSP-01", "Security and Privacy Policy and Procedures"),
    ("DSP-07", "Data Protection by Design and Default"),
    ("DSP-17", "Sensitive Data Protection"),
    ("GRC-01", "Governance Program Policy and Procedures"),
    ("GRC-03", "Organizational Policy Reviews"),
    ("HRS-01", "Background Screening Policy and Procedures"),
    ("HRS-04", "Employment Termination"),
    ("IAM-01", "Identity and Access Management Policy and Procedures"),
    ("IAM-03", "Identity Inventory"),
    ("IAM-08", "User Access Provisioning"),
    ("IAM-12", "Safeguard Logs Integrity"),
    ("IAM-14", "Strong Authentication"),
    ("IPY-01", "Interoperability and Portability Policy and Procedures"),
    ("IVS-01", "Infrastructure and Virtualization Security Policy and Procedures"),
    ("IVS-03", "Network Security"),
    ("IVS-04", "OS Hardening and Base Controls"),
    ("LOG-01", "Logging and Monitoring Policy and Procedures"),
    ("LOG-05", "Audit Logs Monitoring and Response"),
    ("LOG-13", "Failures and Anomalies Reporting"),
    ("SEF-01", "Security Incident Management Policy and Procedures"),
    ("SEF-03", "Incident Response Plans"),
    ("SEF-06", "Incident Response Metrics"),
    ("STA-01", "SSRM (Shared Security Responsibility Model) Policy and Procedures"),
    ("STA-06", "Supply Chain Agreements"),
    ("STA-09", "Third-Party Assessment"),
    ("TVM-01", "Threat and Vulnerability Management Policy and Procedures"),
    ("TVM-02", "Malware Protection"),
    ("TVM-08", "Vulnerability Remediation"),
    ("TVM-10", "Detection Updates"),
    ("UEM-01", "Endpoint Devices Policy and Procedures"),
    ("UEM-09", "Endpoint Detection and Response"),
]


def _db_path() -> str:
    d = os.environ.get("XORCISM_DB_DIR") or r"C:/Users/jerom/XORCISM_databases"
    return os.path.join(d, "XORCISM.db")


def _ensure_vocab(cur: sqlite3.Cursor, name: str) -> int:
    cols = {r[1] for r in cur.execute("PRAGMA table_info(VOCABULARY)").fetchall()}
    namecol = "VocabularyName" if "VocabularyName" in cols else ("Name" if "Name" in cols else None)
    if namecol:
        row = cur.execute(f"SELECT VocabularyID FROM VOCABULARY WHERE {namecol}=?", (name,)).fetchone()
        if row:
            return int(row[0])
    nid = (cur.execute("SELECT COALESCE(MAX(VocabularyID),0) FROM VOCABULARY").fetchone()[0] or 0) + 1
    rec = {"VocabularyID": nid}
    if namecol:
        rec[namecol] = name
    if "VocabularyGUID" in cols:
        rec["VocabularyGUID"] = str(uuid.uuid4())
    keys = list(rec)
    cur.execute(f"INSERT INTO VOCABULARY ({','.join(keys)}) VALUES ({','.join('?'*len(keys))})", [rec[k] for k in keys])
    return nid


def _from_excel(path: str):
    """Parse the official CSA CCM spreadsheet → list of (control_id, title, specification)."""
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    seen = set()
    import re
    cid_re = re.compile(r"^[A-Z&]{2,4}-\d{2}$")
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            # find a cell that looks like a CCM control id
            cid = next((c for c in cells if cid_re.match(c)), "")
            if not cid or cid in seen:
                continue
            idx = cells.index(cid)
            # title/specification = the next non-empty cells after the id
            rest = [c for c in cells[idx + 1:] if c]
            title = rest[0] if rest else cid
            spec = rest[1] if len(rest) > 1 else ""
            seen.add(cid)
            out.append((cid, title[:300], spec[:4000]))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Import CSA CCM into XORCISM.CONTROL")
    ap.add_argument("--file", help="official CCM .xlsx (full 197 controls)")
    a = ap.parse_args()

    if a.file:
        controls = _from_excel(a.file)
    else:
        controls = [(cid, title, "") for cid, title in SEED]
    if not controls:
        print("[ccm] no controls parsed"); return 1

    con = sqlite3.connect(_db_path()); con.execute("PRAGMA busy_timeout=15000"); cur = con.cursor()
    now = datetime.now(timezone.utc).isoformat()
    vid = _ensure_vocab(cur, VOCAB)
    ccols = {r[1] for r in cur.execute("PRAGMA table_info(CONTROL)").fetchall()}
    cur.execute("DELETE FROM CONTROL WHERE VocabularyID=?", (vid,))  # idempotent
    next_id = (cur.execute("SELECT COALESCE(MAX(ControlID),0) FROM CONTROL").fetchone()[0] or 0) + 1

    n = 0
    for cid, title, spec in controls:
        dom = cid.split("-")[0]
        rec = {
            "ControlID": next_id, "ControlGUID": str(uuid.uuid4()),
            "ControlName": f"{cid} {title}".strip(),
            "ControlDescription": f"CSA CCM — {DOMAINS.get(dom, dom)}",
            "VocabularyID": vid, "CIS": cid, "Statement": spec or None,
            "CreatedDate": now, "ValidFromDate": now[:10], "isEncrypted": 0,
        }
        keys = [k for k in rec if k in ccols]
        cur.execute(f"INSERT INTO CONTROL ({','.join(keys)}) VALUES ({','.join('?'*len(keys))})", [rec[k] for k in keys])
        next_id += 1; n += 1
    con.commit(); con.close()
    src = f"{a.file}" if a.file else "embedded seed"
    print(f"[ccm] VocabularyID={vid}: {n} controls across {len({c.split('-')[0] for c,_,_ in controls})} domains ({src}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
