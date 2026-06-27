"""import_dasf.py - import the Databricks AI Security Framework (DASF) into XORCISM.

The DASF (Databricks) is an AI security framework that maps the risks across the AI/ML lifecycle to
recommended mitigation controls. This importer loads the published workbook into XORCISM in three
places, each to its natural home:

  * 73 mitigation controls -> XORCISM.CONTROL, VOCABULARY "Databricks DASF" (a selectable control
    framework; CIS = "DASF N", description = control type / AI-lifecycle step / risks mitigated).
  * 97 AI-lifecycle risks   -> XTHREAT.DASFRISK + DASFCOMPONENT (the AI risk map, mirroring the
    SAIF importer; per-risk component, mitigating control ids, and deployment-model applicability
    - Predictive ML / RAG / Fine-tuned / Pre-trained / Foundational LLMs).
  * 46 third-party AI-security tools -> XORCISM.TOOL (Category "AI Security", with URL) - the DASF
    third-party tooling catalogue (informational; Databricks states this is not an endorsement).

Data: a committed snapshot at importers/data/dasf.json (parsed from the workbook). Pass
`--xlsx <path>` to re-parse a freshly downloaded export
(https://docs.google.com/spreadsheets/d/<id>/export?format=xlsx) and refresh the snapshot.
Idempotent: controls delete-then-insert by VocabularyID; risks/components upsert by Name; tools
upsert by ToolName (tenant 3). DB dir = XORCISM_DB_DIR env or the default.

    python xorcism_python/importers/import_dasf.py
    python xorcism_python/importers/import_dasf.py --xlsx dasf_export.xlsx
"""
from __future__ import annotations

import json
import os
import re
import sqlite3
import sys
import uuid
from datetime import datetime, timezone

VOCAB = "Databricks DASF"
DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "dasf.json")
SOURCE = "https://www.databricks.com/resources/whitepaper/databricks-ai-security-framework-dasf"
TOOL_TENANT = 3


def _db(n: str) -> str:
    d = os.environ.get("XORCISM_DB_DIR") or r"C:/Users/jerom/XORCISM_databases"
    return os.path.join(d, f"{n}.db")


def _cols(cur: sqlite3.Cursor, t: str) -> set:
    return {r[1] for r in cur.execute(f'PRAGMA table_info("{t}")').fetchall()}


def _ins(cur: sqlite3.Cursor, t: str, rec: dict, present: set) -> None:
    keys = [k for k in rec if k in present]
    cur.execute(f"INSERT INTO {t} ({','.join(keys)}) VALUES ({','.join('?'*len(keys))})", [rec[k] for k in keys])


def _ensure_vocab(cur: sqlite3.Cursor, name: str, ref: str, desc: str) -> int:
    cols = _cols(cur, "VOCABULARY")
    nc = "VocabularyName" if "VocabularyName" in cols else "Name"
    now = datetime.now(timezone.utc).isoformat()
    row = cur.execute(f"SELECT VocabularyID FROM VOCABULARY WHERE {nc}=?", (name,)).fetchone()
    if row:
        vid = int(row[0])
        for col, val in (("VocabularyReference", ref), ("VocabularyDescription", desc)):
            if val and col in cols:
                cur.execute(f"UPDATE VOCABULARY SET {col}=? WHERE VocabularyID=?", (val, vid))
        return vid
    vid = (cur.execute("SELECT COALESCE(MAX(VocabularyID),0) FROM VOCABULARY").fetchone()[0] or 0) + 1
    _ins(cur, "VOCABULARY", {"VocabularyID": vid, "VocabularyGUID": str(uuid.uuid4()), "CreatedDate": now,
                             nc: name, "VocabularyReference": ref, "VocabularyDescription": desc}, cols)
    return vid


def _parse_xlsx(path: str) -> dict:
    import openpyxl
    deploy = ["Predictive ML Models", "RAG - LLMs", "Fine-tuned LLMs", "Pre-trained LLMs", "Foundational LLMs"]
    dre = re.compile(r"DASF\s*\d+")
    clean = lambda s: re.sub(r"\s+", " ", str(s)).strip().lstrip("√•●? ").strip() if s is not None else ""
    appl = lambda v: bool(v) and str(v).strip().lower() not in ("no", "n", "n/a", "-", "")
    wb = openpyxl.load_workbook(path, data_only=True)
    risks, controls, tools = [], [], []
    for row in wb["AI Lifecycle Risks"].iter_rows(min_row=3, values_only=True):
        rid = clean(row[0])
        if not rid or not re.search(r"\d+\.\d+$", rid):
            continue
        cids = sorted({re.sub(r"\s+", " ", m.group(0)) for m in dre.finditer(str(row[4] or ""))}, key=lambda x: int(re.search(r"\d+", x).group()))
        risks.append({"id": rid, "component": clean(row[1]), "title": clean(row[2]) or rid, "description": clean(row[3]),
                      "controlIds": cids, "deployment": [deploy[i] for i in range(5) if len(row) > 7 + i and appl(row[7 + i])]})
    for row in wb["Databricks Mitigation Controls"].iter_rows(min_row=3, values_only=True):
        cid = clean(row[0])
        if not re.match(r"^DASF\s*\d+$", cid):
            continue
        controls.append({"id": re.sub(r"\s+", " ", cid), "title": clean(row[1]) or cid, "riskIds": clean(row[2]),
                         "description": clean(row[3]), "sharedResponsibility": clean(row[4]),
                         "controlType": clean(row[10]) if len(row) > 10 else "", "lifecycleStep": clean(row[11]) if len(row) > 11 else ""})
    cat = ""
    for row in wb["Third-Party Tools"].iter_rows(min_row=4):
        if row[0].value and clean(row[0].value):
            cat = clean(row[0].value)
        name = clean(row[1].value)
        if not name or name.lower() in ("tool url", "tool name"):
            continue
        tools.append({"category": cat, "name": name, "description": clean(row[2].value if len(row) > 2 else ""),
                      "url": (row[1].hyperlink.target if row[1].hyperlink else "") or ""})
    comps, seen = [], set()
    for r in risks:
        if r["component"] and r["component"] not in seen:
            seen.add(r["component"]); comps.append(r["component"])
    return {"meta": {"title": "Databricks AI Security Framework (DASF)", "publisher": "Databricks", "source": SOURCE,
                     "risks": len(risks), "controls": len(controls), "tools": len(tools), "components": comps},
            "risks": risks, "controls": controls, "tools": tools}


def _upsert(cur, table, pk, name, cols, present):
    row = cur.execute(f"SELECT {pk} FROM {table} WHERE Name=?", (name,)).fetchone()
    keys = [k for k in cols if k in present]
    if row:
        cur.execute(f"UPDATE {table} SET {', '.join(k+'=?' for k in keys)} WHERE {pk}=?", [cols[k] for k in keys] + [row[0]])
        return 0
    nid = (cur.execute(f"SELECT COALESCE(MAX({pk}),0) FROM {table}").fetchone()[0] or 0) + 1
    allk = [pk, "Name"] + keys
    cur.execute(f"INSERT INTO {table} ({','.join(allk)}) VALUES ({','.join('?'*len(allk))})", [nid, name] + [cols[k] for k in keys])
    return 1


def main() -> int:
    if "--xlsx" in sys.argv:
        data = _parse_xlsx(sys.argv[sys.argv.index("--xlsx") + 1])
        os.makedirs(os.path.dirname(DATA), exist_ok=True)
        json.dump(data, open(DATA, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"[dasf] re-parsed: {data['meta']['risks']} risks, {data['meta']['controls']} controls, {data['meta']['tools']} tools -> {DATA}")
    data = json.load(open(DATA, encoding="utf-8"))
    now = datetime.now(timezone.utc).isoformat()

    # 1) Controls -> XORCISM.CONTROL "Databricks DASF" + 3) Tools -> XORCISM.TOOL
    xo = sqlite3.connect(_db("XORCISM")); xo.execute("PRAGMA busy_timeout=20000"); cur = xo.cursor()
    vid = _ensure_vocab(cur, VOCAB, SOURCE,
                        "Databricks AI Security Framework (DASF) - 73 AI/ML mitigation controls across the AI lifecycle, mapped to 97 risks. Companion to the DASF risk map in XTHREAT.")
    ccols = _cols(cur, "CONTROL")
    cur.execute("DELETE FROM CONTROL WHERE VocabularyID=?", (vid,))
    nid = (cur.execute("SELECT COALESCE(MAX(ControlID),0) FROM CONTROL").fetchone()[0] or 0) + 1
    nc = 0
    for c in data["controls"]:
        ctx = " / ".join(p for p in (c.get("controlType"), c.get("lifecycleStep")) if p)
        desc = f"Databricks DASF{(' / ' + ctx) if ctx else ''}" + (f" - mitigates: {c['riskIds']}" if c.get("riskIds") else "")
        _ins(cur, "CONTROL", {
            "ControlID": nid, "ControlGUID": str(uuid.uuid4()),
            "ControlName": c["title"][:300], "ControlDescription": desc[:600],
            "VocabularyID": vid, "CIS": c["id"], "Statement": (c.get("description") or c["title"])[:2000],
            "CreatedDate": now, "ValidFromDate": now[:10], "isEncrypted": 0,
        }, ccols)
        nid += 1; nc += 1
    # tools
    tcols = _cols(cur, "TOOL")
    tid = (cur.execute("SELECT COALESCE(MAX(ToolID),0) FROM TOOL").fetchone()[0] or 0) + 1
    nt = 0
    for tool in data["tools"]:
        ex = cur.execute("SELECT ToolID FROM TOOL WHERE ToolName=? AND TenantID IS ?", (tool["name"], TOOL_TENANT)).fetchone()
        descr = (tool.get("description") or "").strip()
        descr = (descr + " " if descr else "") + f"[DASF third-party tool - {tool.get('category','')}; informational, not a Databricks endorsement]"
        rec = {"ToolName": tool["name"][:200], "ToolDescription": descr[:1000], "Category": "AI Security",
               "ToolURL": (tool.get("url") or "")[:400], "TenantID": TOOL_TENANT,
               "CreatedDate": now, "ValidFromDate": now[:10], "isEncrypted": 0}
        if ex:
            sets = [k for k in rec if k in tcols and k != "TenantID"]
            cur.execute(f"UPDATE TOOL SET {', '.join(k+'=?' for k in sets)} WHERE ToolID=?", [rec[k] for k in sets] + [ex[0]])
        else:
            rec["ToolID"] = tid; rec["ToolGUID"] = str(uuid.uuid4()); tid += 1
            _ins(cur, "TOOL", rec, tcols); nt += 1
    xo.commit(); xo.close()

    # 2) Risks -> XTHREAT.DASFRISK + DASFCOMPONENT
    xt = sqlite3.connect(_db("XTHREAT")); xt.execute("PRAGMA busy_timeout=20000"); xc = xt.cursor()
    xc.executescript("""
      CREATE TABLE IF NOT EXISTS DASFCOMPONENT (DasfComponentID INTEGER PRIMARY KEY, Name TEXT UNIQUE, MatrixOrder INTEGER, CreatedDate TEXT);
      CREATE TABLE IF NOT EXISTS DASFRISK (DasfRiskID INTEGER PRIMARY KEY, DasfID TEXT, Name TEXT UNIQUE, Description TEXT,
        Component TEXT, Controls TEXT, Deployment TEXT, MatrixOrder INTEGER, URL TEXT, CreatedDate TEXT);
      CREATE INDEX IF NOT EXISTS ix_dasfrisk_component ON DASFRISK(Component);
    """)
    compcols, riskcols = _cols(xc, "DASFCOMPONENT"), _cols(xc, "DASFRISK")
    for i, comp in enumerate(data["meta"].get("components", []), 1):
        _upsert(xc, "DASFCOMPONENT", "DasfComponentID", comp, {"MatrixOrder": i, "CreatedDate": now}, compcols)
    nr = 0
    for i, r in enumerate(data["risks"], 1):
        nr += _upsert(xc, "DASFRISK", "DasfRiskID", r["title"], {
            "DasfID": r["id"], "Description": r.get("description", ""), "Component": r.get("component", ""),
            "Controls": ", ".join(r.get("controlIds") or []), "Deployment": ", ".join(r.get("deployment") or []),
            "MatrixOrder": i, "URL": SOURCE, "CreatedDate": now}, riskcols)
    xt.commit(); xt.close()

    print(f"[dasf] CONTROL VocabID={vid}: {nc} controls; XTHREAT: {len(data['risks'])} risks ({nr} new); TOOL: +{nt} new AI-security tools (tenant {TOOL_TENANT}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
