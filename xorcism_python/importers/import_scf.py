"""import_scf.py — import the Secure Controls Framework (SCF) into XORCISM.CONTROL.

The SCF (securecontrolsframework.com) is a free, comprehensive cybersecurity & privacy
"metaframework": ~1000+ controls across ~33 domains, each cross-mapped to 100+ statutory,
regulatory and contractual authorities (NIST 800-53, ISO 27001/2, CIS, PCI DSS, SOC 2, GDPR,
CMMC, CCPA, HIPAA…). It is the Rosetta Stone of control frameworks.

Two modes:
  * --file <SCF.xlsx>  : parse the official SCF spreadsheet (openpyxl). Reads the "SCF #" /
                         "SCF Control" / "Secure Controls Framework (SCF) Control Description"
                         columns from the main worksheet, and (best-effort) the per-authority
                         mapping columns → CONTROLMAPPING. (Download from securecontrolsframework.com.)
  * (no file)          : seed the ~33 SCF domains + a representative set of controls (embedded),
                         so the framework is present immediately; run --file for the full catalogue.

Writes CONTROL rows (VocabularyID = the SCF vocab; ControlName="SCF-ID Title", domain in
ControlDescription, Statement=description). Idempotent by VocabularyID. Raw SQL.

    python xorcism_python/importers/import_scf.py [--file SCF.xlsx]
"""
from __future__ import annotations

import argparse
import os
import re
import sqlite3
import uuid
from datetime import datetime, timezone

VOCAB = "SCF"
SRC = "SCF 2026.2"  # CONTROLMAPPING.Source marker (idempotency key for the cross-mappings)
_SCF_ID_RE = re.compile(r"^[A-Z]{3}-\d{2}(?:\.\d+)?$")
# Header substrings that mark SCF-internal columns (NOT external-authority framework mappings):
# metadata, MCR/DSR designations, and the SCF Risk/Threat catalog applicability matrices.
_META_HINTS = ("conformity validation", "evidence request", "possible solutions", "control question",
               "relative control weighting", "pptdf", "function grouping", "scrm focus", "scr-cmm",
               "community derived", "threat summary", "errata", "minimum security requirements",
               "minimum compliance requirements", "discretionary security requirements")


def _clean_hdr(h) -> str:
    return " ".join(str(h if h is not None else "").replace("\n", " ").split())


def _is_authority(hdr: str) -> bool:
    """A column header that names an external framework/authority (vs SCF-internal columns)."""
    lo = hdr.lower()
    if not lo or lo.startswith("scf"):
        return False
    if lo.startswith("risk ") or lo.startswith("threat "):  # SCF Risk / Threat catalog columns
        return False
    return not any(k in lo for k in _META_HINTS)

DOMAINS = {
    "GOV": "Security & Privacy Governance", "AST": "Asset Management",
    "BCD": "Business Continuity & Disaster Recovery", "CAP": "Capacity & Performance Planning",
    "CFG": "Configuration Management", "CHG": "Change Management", "CLD": "Cloud Security",
    "CPL": "Compliance", "CRY": "Cryptographic Protections", "DCH": "Data Classification & Handling",
    "END": "Endpoint Security", "HRS": "Human Resources Security", "IAC": "Identification & Authentication",
    "IAO": "Information Assurance", "IRO": "Incident Response", "MNT": "Maintenance",
    "MDM": "Mobile Device Management", "NET": "Network Security", "PES": "Physical & Environmental Security",
    "PRI": "Privacy", "PRM": "Project & Resource Management", "RSK": "Risk Management",
    "SAT": "Security Awareness & Training", "SEA": "Secure Engineering & Architecture",
    "TDA": "Technology Development & Acquisition", "THR": "Threat Management", "TPM": "Third-Party Management",
    "VPM": "Vulnerability & Patch Management", "WEB": "Web Security", "MON": "Continuous Monitoring",
    "AAT": "Artificial & Autonomous Technologies", "EMB": "Embedded Technology", "OPS": "Security Operations",
}

SEED = [
    ("GOV-01", "Information Security & Privacy Governance Program"),
    ("GOV-02", "Publishing Security & Privacy Documentation"),
    ("GOV-05", "Operationalizing Cybersecurity & Data Protection Practices"),
    ("AST-01", "Asset Governance"), ("AST-02", "Asset Inventories"),
    ("BCD-01", "Business Continuity Management System (BCMS)"), ("BCD-11", "Data Backups"),
    ("CAP-01", "Capacity & Performance Management"),
    ("CFG-01", "Configuration Management Program"), ("CFG-02", "System Hardening Through Baseline Configurations"),
    ("CHG-01", "Change Management Program"),
    ("CLD-01", "Cloud Services"), ("CLD-06", "Multi-Tenant Environments"),
    ("CPL-01", "Statutory, Regulatory & Contractual Compliance"), ("CPL-02", "Cybersecurity & Data Privacy Controls Oversight"),
    ("CRY-01", "Use of Cryptographic Controls"), ("CRY-05", "Encrypting Data At Rest"), ("CRY-09", "Cryptographic Key Management"),
    ("DCH-01", "Data Protection"), ("DCH-02", "Data & Asset Classification"),
    ("END-01", "Endpoint Security"), ("END-04", "Malicious Code Protection (Anti-Malware)"),
    ("HRS-01", "Human Resources Security Management"), ("HRS-04", "Personnel Screening"),
    ("IAC-01", "Identity & Access Management (IAM)"), ("IAC-06", "Multi-Factor Authentication (MFA)"), ("IAC-10", "Authenticator Management"),
    ("IAO-01", "Information Assurance (IA) Operations"),
    ("IRO-01", "Incident Response Operations"), ("IRO-02", "Incident Handling"),
    ("MNT-01", "Maintenance Operations"), ("MDM-01", "Centralized Management Of Mobile Devices"),
    ("NET-01", "Network Security Management"), ("NET-06", "Boundary Protection"),
    ("PES-01", "Physical & Environmental Protections"),
    ("PRI-01", "Privacy Program"), ("PRI-05", "Data Subject Access"),
    ("PRM-01", "Cybersecurity & Data Protection Portfolio Management"),
    ("RSK-01", "Risk Management Program"), ("RSK-04", "Risk Assessment"), ("RSK-06", "Risk Remediation"),
    ("SAT-01", "Security & Privacy-Minded Workforce"),
    ("SEA-01", "Secure Engineering Principles"), ("SEA-02", "Alignment With Enterprise Architecture"),
    ("TDA-01", "Technology Development & Acquisition"), ("TDA-06", "Secure Coding"),
    ("THR-01", "Threat Intelligence Program"), ("THR-03", "Threat Hunting"),
    ("TPM-01", "Third-Party Management"), ("TPM-05", "Third-Party Contract Requirements"),
    ("VPM-01", "Vulnerability & Patch Management Program (VPMP)"), ("VPM-05", "Software & Firmware Patching"), ("VPM-06", "Vulnerability Scanning"),
    ("WEB-01", "Web Security"),
    ("MON-01", "Continuous Monitoring"), ("MON-02", "Centralized Collection of Security Event Logs"),
    ("AAT-01", "Artificial & Autonomous Technologies Governance"), ("AAT-02", "AI & Autonomous Technologies Risk Management"),
    ("OPS-01", "Operations Security"),
]


def _db_path() -> str:
    d = os.environ.get("XORCISM_DB_DIR") or r"C:/Users/jerom/XORCISM_databases"
    return os.path.join(d, "XORCISM.db")


def _ensure_vocab(cur: sqlite3.Cursor, name: str) -> int:
    cols = {r[1] for r in cur.execute("PRAGMA table_info(VOCABULARY)").fetchall()}
    namecol = "VocabularyName" if "VocabularyName" in cols else ("Name" if "Name" in cols else None)
    if namecol:
        row = cur.execute(f"SELECT VocabularyID FROM VOCABULARY WHERE {namecol}=?", (name,)).fetchone()
        if row:
            return int(row[0])
    nid = (cur.execute("SELECT COALESCE(MAX(VocabularyID),0) FROM VOCABULARY").fetchone()[0] or 0) + 1
    rec = {"VocabularyID": nid}
    if namecol:
        rec[namecol] = name
    if "VocabularyGUID" in cols:
        rec["VocabularyGUID"] = str(uuid.uuid4())
    keys = list(rec)
    cur.execute(f"INSERT INTO VOCABULARY ({','.join(keys)}) VALUES ({','.join('?'*len(keys))})", [rec[k] for k in keys])
    return nid


def _from_excel(path: str):
    """Parse the official SCF worksheet → list of control dicts.

    Each = {sid, domain, title, desc, maps:[(framework, ref), …]}. The control columns
    (SCF Domain / SCF Control / SCF # / Control Description) are located by header text;
    every remaining external-authority column (ISO, NIST, PCI, CIS, ATT&CK…) with a value
    yields cross-mappings (cells are newline-separated reference lists).
    """
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = next((w for w in wb.worksheets if re.match(r"SCF\s*\d", w.title or "")), wb.worksheets[0])
    it = ws.iter_rows(values_only=True)
    hdr = [_clean_hdr(h) for h in next(it)]

    def _col(pred, default):
        for i, h in enumerate(hdr):
            if pred(h.lower()):
                return i
        return default
    c_id = _col(lambda h: h == "scf #" or h.startswith("scf #"), 2)
    c_dom = _col(lambda h: h.startswith("scf domain"), 0)
    c_ttl = _col(lambda h: h == "scf control", 1)
    c_desc = _col(lambda h: "control description" in h, 3)
    auth_cols = [i for i in range(c_desc + 1, len(hdr)) if _is_authority(hdr[i])]

    out, seen = [], set()
    for row in it:
        cells = ["" if c is None else str(c) for c in row]
        if len(cells) <= c_id:
            continue
        sid = cells[c_id].strip()
        if not _SCF_ID_RE.match(sid) or sid in seen:
            continue
        seen.add(sid)
        get = lambda i: cells[i].strip() if i < len(cells) else ""
        maps = []
        for i in auth_cols:
            if i >= len(cells) or not cells[i].strip():
                continue
            for ref in re.split(r"[\r\n]+", cells[i]):
                ref = ref.strip()
                if ref:
                    maps.append((hdr[i], ref[:120]))
        out.append({"sid": sid, "domain": get(c_dom), "title": get(c_ttl)[:300] or sid,
                    "desc": get(c_desc)[:4000], "maps": maps})
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Import the Secure Controls Framework into XORCISM.CONTROL")
    ap.add_argument("--file", help="official SCF .xlsx (full catalogue + cross-mappings)")
    ap.add_argument("--db-dir", help="directory holding XORCISM.db (default: $XORCISM_DB_DIR)")
    a = ap.parse_args()

    if a.file:
        records = _from_excel(a.file)
    else:  # no file → the embedded representative seed (framework present immediately, no mappings)
        records = [{"sid": sid, "domain": DOMAINS.get(sid.split("-")[0], ""), "title": title, "desc": "", "maps": []}
                   for sid, title in SEED]
    if not records:
        print("[scf] no controls parsed"); return 1

    dbp = os.path.join(a.db_dir, "XORCISM.db") if a.db_dir else _db_path()
    con = sqlite3.connect(dbp); con.execute("PRAGMA busy_timeout=30000"); cur = con.cursor()
    now = datetime.now(timezone.utc).isoformat()
    vid = _ensure_vocab(cur, VOCAB)
    ccols = {r[1] for r in cur.execute("PRAGMA table_info(CONTROL)").fetchall()}
    have_map = bool(cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='CONTROLMAPPING'").fetchone())

    # Idempotent refresh: drop the previous SCF controls (by vocab) and its cross-mappings (by source).
    cur.execute("DELETE FROM CONTROL WHERE VocabularyID=?", (vid,))
    if have_map:
        cur.execute("DELETE FROM CONTROLMAPPING WHERE Source=?", (SRC,))
    next_id = (cur.execute("SELECT COALESCE(MAX(ControlID),0) FROM CONTROL").fetchone()[0] or 0) + 1
    next_map = ((cur.execute("SELECT COALESCE(MAX(MappingID),0) FROM CONTROLMAPPING").fetchone()[0] or 0) + 1) if have_map else 0

    domains, map_rows = set(), []
    for r in records:
        dom = r["sid"].split("-")[0]; domains.add(dom)
        rec = {
            "ControlID": next_id, "ControlGUID": f"scf-{r['sid']}",
            "ControlName": f"{r['sid']} {r['title']}".strip(),
            "ControlDescription": r["domain"] or f"SCF — {DOMAINS.get(dom, dom)}",
            "VocabularyID": vid, "Statement": r["desc"] or None,
            "CreatedDate": now, "ValidFromDate": now[:10], "isEncrypted": 0,
        }
        keys = [k for k in rec if k in ccols]
        cur.execute(f"INSERT INTO CONTROL ({','.join(keys)}) VALUES ({','.join('?'*len(keys))})", [rec[k] for k in keys])
        if have_map:
            for fw, ref in r["maps"]:
                map_rows.append((next_map, str(uuid.uuid4()), next_id, fw, ref, "mapped-to", SRC, now))
                next_map += 1
        next_id += 1
    if have_map and map_rows:
        cur.executemany(
            "INSERT INTO CONTROLMAPPING (MappingID, MappingGUID, ControlID, Framework, ExternalID, "
            "Relationship, Source, CreatedDate) VALUES (?,?,?,?,?,?,?,?)", map_rows)
    try:  # keep the SCF FRAMEWORK row's version current
        cur.execute("UPDATE FRAMEWORK SET FrameworkVersion=? WHERE FrameworkName LIKE '%Secure Controls Framework%' "
                    "OR FrameworkName LIKE '%SCF%'", ("2026.2",))
    except Exception:  # noqa: BLE001
        pass
    con.commit(); con.close()
    src = a.file if a.file else "embedded seed"
    print(f"[scf] VocabularyID={vid}: {len(records)} controls / {len(domains)} domains, "
          f"{len(map_rows)} cross-mappings across {len({fw for _,_,_,fw,_,_,_,_ in map_rows})} authorities ({src}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
