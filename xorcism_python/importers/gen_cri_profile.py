"""gen_cri_profile.py - parse the official CRI Profile v2.2 xlsx into the bundled cri_profile_v2.2.json.

The Cyber Risk Institute (CRI) Profile is the financial-sector cybersecurity benchmark: NIST CSF 2.0
structure (Govern/Identify/Protect/Detect/Respond/Recover) extended with CRI "Diagnostic Statements"
(the testable controls), impact-tier applicability (Tier-1..4) and a NIST CSF v2 crosswalk.

Reads the "CRI Profile v2.2 Structure" sheet (the canonical hierarchy) + the "Assessment" sheet (short
Diagnostic-Statement names), and emits a stable JSON the runtime importer consumes (so the importer
needs neither openpyxl nor the xlsx). ASCII-only output.

    python xorcism_python/importers/gen_cri_profile.py [path-to-xlsx]
"""
from __future__ import annotations

import json
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(HERE, "..", "..", "resources", "CRI-Profile-ver.-2.2.2026-04-27.xlsx")
OUT = os.path.join(HERE, "cri_profile_v2.2.json")


def clean(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").replace("\xa0", " ")).strip()


def main() -> int:
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    # Short Diagnostic-Statement names from the Assessment sheet: Profile Id -> name.
    short = {}
    aws = wb["CRI Profile v2.2 Assessment"]
    for row in aws.iter_rows(min_row=5, values_only=True):
        pid = clean(row[2]) if len(row) > 2 else ""   # Profile Id
        nm = clean(row[3]) if len(row) > 3 else ""     # Diagnostic Statement Name
        if pid and nm:
            short[pid] = nm

    ws = wb["CRI Profile v2.2 Structure"]
    controls = []
    func, cat, sub = "", "", ""
    funcs, cats = set(), set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 5:
            continue
        level = clean(row[1])
        pid = clean(row[2])
        path = clean(row[3])
        stmt = clean(row[4])
        if not pid:
            continue
        if level == "F":
            func = path or pid
            funcs.add(pid)
            continue
        if level == "C":
            cat = path or pid
            cats.add(pid)
            continue
        if level == "S":
            sub = pid
            continue
        if level != "DS":
            continue
        # Diagnostic Statement = a testable control. Strip the leading "ID: " from the statement text.
        text = re.sub(r"^" + re.escape(pid) + r"\s*:\s*", "", stmt)
        tiers = [t for t, col in zip((1, 2, 3, 4), row[5:9]) if str(col or "").strip().lower() == "yes"]
        nist = clean(re.sub(r"\(CRI Modified\)", "", str(row[9] or "")))
        fsmap = clean(row[10]) if len(row) > 10 else ""
        controls.append({
            "id": pid,
            "name": short.get(pid, ""),
            "function": func,
            "category": path or cat,
            "statement": text[:6000],
            "tiers": tiers,
            "nist": nist,
            "fsMapping": fsmap[:2000],
        })

    out = {
        "framework": "CRI Profile",
        "version": "2.2",
        "released": "2026-04-27",
        "source": "Cyber Risk Institute (CRI) Profile v2.2",
        "functions": sorted(funcs),
        "categories": sorted(cats),
        "controls": controls,
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=True, indent=1)
    print(f"controls={len(controls)} functions={len(funcs)} categories={len(cats)} with-nist={sum(1 for c in controls if c['nist'])} -> {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
